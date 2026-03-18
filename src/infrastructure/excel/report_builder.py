from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.domain.report.entities import WordStats

_HEADERS = ["словоформа", "кол-во во всём документе", "кол-во в каждой строке"]
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)


class ExcelReportBuilder:
    """Builds an .xlsx report from word frequency statistics."""

    def build(
        self,
        stats: dict[str, WordStats],
        total_lines: int,
        output_path: str,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Conclusion"

        # Headers
        ws.append(_HEADERS)
        for col, _ in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for word_stats in stats.values():
            ws.append(
                [
                    word_stats.word,
                    word_stats.total,
                    word_stats.per_line_str(total_lines),
                ]
            )

        # Auto-fit column widths (approximate)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 50

        wb.save(output_path)
