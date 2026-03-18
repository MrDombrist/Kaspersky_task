import os
import uuid

import pytest
from openpyxl import load_workbook

from src.application.report.use_cases import SyncProcessFileUseCase
from src.domain.report.entities import JobStatus, ReportJob
from src.infrastructure.excel.report_builder import ExcelReportBuilder


@pytest.fixture
def processor(mock_normalizer):
    return SyncProcessFileUseCase(
        report_builder=ExcelReportBuilder(),
        normalizer=mock_normalizer,
    )


@pytest.fixture
def text_file(tmp_path):
    """Write a small text file and return its path."""
    content = "жители города\nжителем и горожанин\nгород и жители\n"
    p = tmp_path / "input.txt"
    p.write_text(content, encoding="utf-8")
    return str(p)


@pytest.fixture
def result_path(tmp_path):
    return str(tmp_path / "result.xlsx")


class TestSyncProcessFileUseCase:
    def test_successful_processing_sets_done(self, processor, text_file, result_path):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=text_file,
            result_path=result_path,
        )
        processor.process(job)
        assert job.status == JobStatus.DONE

    def test_result_file_created(self, processor, text_file, result_path):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=text_file,
            result_path=result_path,
        )
        processor.process(job)
        assert os.path.exists(result_path)

    def test_input_file_deleted_after_processing(
        self, processor, text_file, result_path
    ):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=text_file,
            result_path=result_path,
        )
        processor.process(job)
        assert not os.path.exists(text_file)

    def test_total_lines_set_correctly(self, processor, text_file, result_path):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=text_file,
            result_path=result_path,
        )
        processor.process(job)
        assert job.total_lines == 3

    def test_xlsx_contains_expected_lemmas(self, processor, text_file, result_path):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=text_file,
            result_path=result_path,
        )
        processor.process(job)

        wb = load_workbook(result_path)
        ws = wb.active
        words = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "житель" in words
        assert "город" in words

    def test_failed_job_on_missing_file(self, processor, result_path):
        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path="/non/existent/file.txt",
            result_path=result_path,
        )
        processor.process(job)
        assert job.status == JobStatus.FAILED
        assert job.error is not None

    def test_empty_file_produces_header_only(self, processor, tmp_path, result_path):
        empty = tmp_path / "empty.txt"
        empty.write_text("", encoding="utf-8")

        job = ReportJob(
            job_id=uuid.uuid4(),
            file_path=str(empty),
            result_path=result_path,
        )
        processor.process(job)
        assert job.status == JobStatus.DONE

        wb = load_workbook(result_path)
        assert wb.active.max_row == 1  # header only
