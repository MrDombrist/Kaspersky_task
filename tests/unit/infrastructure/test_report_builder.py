import os

import pytest
from openpyxl import load_workbook

from src.domain.report.entities import WordStats
from src.infrastructure.excel.report_builder import ExcelReportBuilder


@pytest.fixture
def builder():
    return ExcelReportBuilder()


@pytest.fixture
def sample_stats() -> dict[str, WordStats]:
    ws_житель = WordStats(word="житель")
    ws_житель.add(0, 2)
    ws_житель.add(2, 1)

    ws_город = WordStats(word="город")
    ws_город.add(1, 5)

    return {"житель": ws_житель, "город": ws_город}


class TestExcelReportBuilder:
    def test_creates_file(self, builder, sample_stats, tmp_path):
        out = str(tmp_path / "report.xlsx")
        builder.build(sample_stats, total_lines=3, output_path=out)
        assert os.path.exists(out)

    def test_header_row(self, builder, sample_stats, tmp_path):
        out = str(tmp_path / "report.xlsx")
        builder.build(sample_stats, total_lines=3, output_path=out)

        wb = load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        assert headers == [
            "словоформа",
            "кол-во во всём документе",
            "кол-во в каждой строке",
        ]

    def test_data_rows_correct_values(self, builder, sample_stats, tmp_path):
        out = str(tmp_path / "report.xlsx")
        builder.build(sample_stats, total_lines=3, output_path=out)

        wb = load_workbook(out)
        ws = wb.active
        data = {
            ws.cell(r, 1).value: (ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(2, ws.max_row + 1)
        }

        assert data["житель"] == (3, "2,0,1")
        assert data["город"] == (5, "0,5,0")

    def test_sheet_title(self, builder, sample_stats, tmp_path):
        out = str(tmp_path / "report.xlsx")
        builder.build(sample_stats, total_lines=3, output_path=out)
        wb = load_workbook(out)
        assert wb.active.title == "Conclusion"

    def test_empty_stats_only_header(self, builder, tmp_path):
        out = str(tmp_path / "empty.xlsx")
        builder.build({}, total_lines=0, output_path=out)
        wb = load_workbook(out)
        assert wb.active.max_row == 1

    def test_single_line_document(self, builder, tmp_path):
        ws = WordStats(word="слово")
        ws.add(0, 3)
        out = str(tmp_path / "single.xlsx")
        builder.build({"слово": ws}, total_lines=1, output_path=out)

        wb = load_workbook(out)
        row2 = [wb.active.cell(2, c).value for c in range(1, 4)]
        assert row2 == ["слово", 3, "3"]
